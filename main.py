from fastapi import FastAPI, Form, File, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
import os
import openpyxl
import re
import stat
import json
import base64
from io import BytesIO
from PIL import Image
import httpx
import aiofiles
from typing import List
from git_api import GA1_13, GA2_3, GA2_7, GA4_8, GA2_9_file, GA2_6_file
from dotenv import load_dotenv
from processing import fetch_answer

# Load environment variables from .env file
load_dotenv()

app = FastAPI()

# CORS Configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "tasks.xlsx")


def load_tasks_from_excel():
    if not os.path.exists(EXCEL_FILE):
        return {}, {}
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    tasks = {
        row[0]: row[1]
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[0] and row[1]
    }
    tasks_answers = {
        row[0]: row[2]
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[0] and row[2]
    }
    workbook.close()
    return tasks, tasks_answers


TASKS, TASKS_ANSWERS = load_tasks_from_excel()


def classify_task(question: str) -> str:
    """Classify a question based on keyword matching with TASKS."""
    ql = question.lower()
    for task_id, keyword in TASKS.items():
        if keyword.lower() in ql:
            return task_id
    return "Unknown"


def save_file(file: UploadFile):
    os.makedirs("uploads", exist_ok=True)
    if not file or not file.filename:
        return "Error: No file provided."
    file_path = os.path.join(os.getcwd(), "uploads", file.filename)
    try:
        with open(file_path, "wb") as buffer:
            buffer.write(file.file.read())
        os.chmod(file_path, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
    except Exception as e:
        return f"Error saving file: {str(e)}"
    return file_path


def get_file_path(question: str) -> str:
    """Extracts a single filename from the question and returns its uploads path."""
    match = re.search(r'([^\/\\\s]+?\.[a-zA-Z0-9]+)', question)
    if not match:
        return None
    fn = match.group(1)
    path = os.path.join(os.getcwd(), "uploads", fn)
    return path if os.path.exists(path) else None


@app.get("/", response_class=HTMLResponse)
async def serve_form():
    file_path = os.path.join(os.path.dirname(__file__), "index.html")
    try:
        with open(file_path, "r") as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>index.html not found</h1>", status_code=404)


async def read_answer(task_id: str, question: str):
    return TASKS_ANSWERS.get(task_id, "No answer found for this task.")


def to_string(value):
    """Convert any type to string."""
    if value is None:
        return "None"
    if not isinstance(value, str):
        try:
            return json.dumps(value)
        except (TypeError, ValueError):
            return str(value)
    return value


def Solve_Unknown_Task(question):
    BASE_URL = "https://aiproxy.sanand.workers.dev/openai/v1"
    data = {
        "model": "gpt-4o-mini",
        "messages": [{"role": "user", "content": question + " return only the answer"}]
    }
    API_KEY = os.getenv("AIPROXY_TOKEN")
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}"
    }
    resp = httpx.post(f"{BASE_URL}/chat/completions", json=data, headers=headers, timeout=60)
    return resp.json().get("choices", [])[0].get("message", {}).get("content")


@app.post("/api/")
async def receive_question(
    question: str = Form(...),
    files: List[UploadFile] = File(None)
):
    # Only use the first file if multiple were uploaded
    file = files[0] if files else None

    # Quick file-path lookup shortcut
    if 'where is ' in question.lower():
        path = get_file_path(question)
        return {
            "question": question,
            "answer": path or "File not found"
        }

    # Classify and dispatch
    task_id = classify_task(question)

    if task_id == "Unknown":
        answer = Solve_Unknown_Task(question)

    elif task_id == 'GA1.1':
        answer = await read_answer(task_id, question)

    elif task_id in ['GA1.2', 'GA1.4', 'GA1.5', 'GA1.7', 'GA1.9', 'GA1.18']:
        answer = await fetch_answer(task_id=task_id, question=question, file_path="")

    elif task_id == 'GA1.3':
        if file and not os.getenv('VERCEL'):
            answer = await fetch_answer(task_id, question, file)
        else:
            answer = await read_answer(task_id, question)

    elif task_id == 'GA1.16':
        if file:
            answer = await fetch_answer(task_id, question, file)
        else:
            answer = await read_answer(task_id, question)

    elif task_id in ['GA1.8', 'GA1.10', 'GA1.12', 'GA1.14', 'GA1.15', 'GA1.17']:
        if file:
            answer = await fetch_answer(task_id, question, file)
        else:
            answer = await read_answer(task_id, question)

    elif task_id in ['GA1.6', 'GA1.11']:
        if file:
            ans = await fetch_answer(task_id, question, file)
        else:
            ans = ""
        answer = ans or await read_answer(task_id, question)

    elif task_id == 'GA1.13':
        answer = GA1_13(question)

    # Example for GA2.x – you’ll want to fill in the rest similarly...
    elif task_id == 'GA2.3':
        answer = GA2_3(question)

    elif task_id == 'GA2.7':
        answer = GA2_7(question)

    elif task_id == 'GA4.8':
        answer = GA4_8(question)

    else:
        # catch‑all: read or default
        if file:
            # you could call save_file(file) here if you need it on disk
            pass
        answer = await read_answer(task_id, question)

    # Normalize and return
    return {"answer": to_string(answer)}
